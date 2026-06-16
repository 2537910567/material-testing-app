import sys
import os
import tempfile
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from app.report.report_generator import generate_testing_plan, _merge_section_cells
import openpyxl


class TestGenerateTestingPlan:
    def test_minimal_generation(self, tmpdir, sample_plan):
        output = str(tmpdir / "test_output.xlsx")
        result = generate_testing_plan(
            output_path=output,
            project_info={"project_name": "测试项目"},
            testing_plan=sample_plan,
            project_name="测试项目"
        )
        assert os.path.exists(result)
        assert result.endswith(".xlsx")

    def test_output_has_expected_sheets(self, tmpdir, sample_plan):
        """V4.7: 2 sheets without construction_layers, 3 sheets with them"""
        output = str(tmpdir / "test_output.xlsx")
        result = generate_testing_plan(
            output_path=output,
            project_info={"project_name": "Test"},
            testing_plan=sample_plan,
            project_name="Test"
        )
        wb = openpyxl.load_workbook(result)
        assert "封面" in wb.sheetnames
        assert "按路段送检计划" in wb.sheetnames
        # V4.7: "按GB55032分类" deleted; "施工检测工序流程" only with construction_layers
        assert "按GB55032分类" not in wb.sheetnames

    def test_output_with_construction_layers(self, tmpdir, sample_plan):
        """V4.7: construction_layers triggers third sheet"""
        output = str(tmpdir / "test_with_layers.xlsx")
        layers = [{
            "section_name": "K0+100~K0+500",
            "road_orientation": "双侧",
            "step": 1,
            "layer_name": "基层",
            "thickness": "200mm",
            "construction_process": "摊铺→碾压",
            "materials": [{"name": "碎石", "spec": "0~31.5mm"}],
            "tests": [{"test_item": "压实度", "test_param": "", "timing": "", "frequency": "", "standard": ""}],
        }]
        result = generate_testing_plan(
            output_path=output,
            project_info={"project_name": "Test"},
            testing_plan=sample_plan,
            project_name="Test",
            construction_layers=layers,
        )
        wb = openpyxl.load_workbook(result)
        assert "封面" in wb.sheetnames
        assert "按路段送检计划" in wb.sheetnames
        assert "施工检测工序流程" in wb.sheetnames

    def test_empty_plan_does_not_crash(self, tmpdir):
        output = str(tmpdir / "empty.xlsx")
        result = generate_testing_plan(
            output_path=output,
            project_info={},
            testing_plan=[],
            project_name="Empty"
        )
        assert os.path.exists(result)


class TestMergeSectionCells:
    def test_single_row_no_merge(self, tmpdir):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(3, 2, value="A")
        ws.cell(3, 3, value="双侧")
        # Should not raise
        _merge_section_cells(ws, 3, 3)

    def test_empty_range_no_merge(self, tmpdir):
        wb = openpyxl.Workbook()
        ws = wb.active
        # start > end
        _merge_section_cells(ws, 5, 3)
