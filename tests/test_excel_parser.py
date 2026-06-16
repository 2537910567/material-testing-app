"""Tests for Excel (.xlsx) parser — V4.9"""
import pytest
import sys
import os
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from app.engine.excel_parser import extract_excel_content


class TestExcelParser:
    """Test Excel spreadsheet content extraction."""

    def test_missing_file(self):
        """Non-existent file returns error text."""
        result = extract_excel_content("/nonexistent/file.xlsx")
        assert "text" in result
        assert "ERROR" in result["text"]

    def test_empty_workbook(self):
        """Empty workbook returns minimal result."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
            wb = openpyxl.Workbook()
            wb.save(tmp_path)

        try:
            result = extract_excel_content(tmp_path)
            assert "text" in result
            assert result["pages"] == 1
        finally:
            os.unlink(tmp_path)

    def test_single_sheet(self):
        """Single sheet with data is extracted."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "材料清单"
            ws.cell(1, 1, "材料名称")
            ws.cell(2, 1, "水泥")
            wb.save(tmp_path)

        try:
            result = extract_excel_content(tmp_path)
            assert "材料清单" in result["text"]
            assert "水泥" in result["text"]
        finally:
            os.unlink(tmp_path)

    def test_multi_sheet(self):
        """Multiple sheets are all extracted."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
            wb = openpyxl.Workbook()
            ws1 = wb.active
            ws1.title = "Sheet1"
            ws1.cell(1, 1, "A1")
            ws2 = wb.create_sheet("Sheet2")
            ws2.cell(1, 1, "B1")
            wb.save(tmp_path)

        try:
            result = extract_excel_content(tmp_path)
            assert len(result["sheets"]) >= 2
        finally:
            os.unlink(tmp_path)
