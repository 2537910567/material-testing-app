"""
Excel (.xlsx) 表格解析模块
提取所有 sheet 的单元格文字，供 AI 文本分析使用。

V4.9: 新增文件格式支持。
"""

from pathlib import Path
from typing import List, Dict
import re


def extract_excel_content(file_path: str) -> dict:
    """
    解析 .xlsx 文件，提取所有 sheet 的文字和表格。

    Args:
        file_path: .xlsx 文件路径

    Returns:
        {"text": str, "sheets": List[dict], "pages": 1}
        - text: 所有 sheet 合并的文字
        - sheets: [{name, text, tables}] 每个 sheet 的详细信息
        - pages: 始终为 1
    """
    try:
        import openpyxl
    except ImportError:
        return {"text": "[ERROR: openpyxl not installed]", "sheets": [], "pages": 1}

    path = Path(file_path)
    if not path.exists():
        return {"text": f"[ERROR: File not found: {file_path}]", "sheets": [], "pages": 1}

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        return {"text": f"[ERROR: Failed to open {path.name}: {e}]", "sheets": [], "pages": 1}

    all_text_parts = []
    sheets_info = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_text_parts = []
        tables_md = []

        rows = []
        for row in ws.iter_rows(values_only=True):
            row_values = [str(c).strip() if c is not None else "" for c in row]
            if any(v for v in row_values):
                while row_values and not row_values[-1]:
                    row_values.pop()
                if row_values:
                    rows.append(row_values)

        if not rows:
            continue

        sheet_text_parts.append(f"=== Sheet: {sheet_name} ===")
        if len(rows) >= 2:
            max_cols = max(len(r) for r in rows)
            md_lines = []
            header = rows[0]
            while len(header) < max_cols:
                header.append("")
            md_lines.append("| " + " | ".join(header) + " |")
            md_lines.append("|" + "|".join(["---"] * max_cols) + "|")
            for row in rows[1:]:
                while len(row) < max_cols:
                    row.append("")
                md_lines.append("| " + " | ".join(row) + " |")
            tables_md.append("\n".join(md_lines))

        for row in rows:
            sheet_text_parts.append(" | ".join(row))

        all_text_parts.append("\n".join(sheet_text_parts))
        sheets_info.append({
            "name": sheet_name,
            "text": "\n".join(sheet_text_parts),
            "tables": tables_md,
        })

    wb.close()

    full_text = "\n\n".join(all_text_parts)
    full_text = re.sub(r'\n{3,}', '\n\n', full_text)

    return {
        "text": full_text,
        "sheets": sheets_info,
        "pages": 1,
    }
