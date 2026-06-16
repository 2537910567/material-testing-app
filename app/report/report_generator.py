"""
Excel 报告生成模块 V3 — 基于 GB55032-2022 的送检计划模板
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional

# ===== 样式 =====
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True)
HEADER_FONT = Font(name="微软雅黑", size=9, bold=True, color="FFFFFF")
SECTION_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
DATA_FONT = Font(name="微软雅黑", size=8)

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FILL = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
ROW_ODD_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
ROW_EVEN_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
ALT_SECTION_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

# ===== V3 输出列（16列） =====
GB55032_COLUMNS = [
    ("序号", 5),
    ("路段(桩号)", 16),
    ("路幅", 8),
    ("车道", 6),
    ("分部工程", 10),
    ("子分部工程", 12),
    ("分项工程", 22),
    ("材料名称", 14),
    ("规格型号", 10),
    ("检测项目", 16),
    ("检测参数", 20),
    ("检测标准", 22),
    ("取样方法", 14),
    ("送检类型", 10),
    ("检验批/取样频率", 26),
    ("计划批次", 8),
    ("备注", 14),
]

# ===== V6.0 工序流程列（用于 Sheet 3 — 12列4级层级） =====
PROCESS_FLOW_COLS = [
    ("路段(桩号)", 16),
    ("路幅", 8),
    ("施工层序", 8),
    ("施工层/工序", 28),
    ("施工要点", 22),
    ("材料名称", 14),
    ("规格型号", 12),
    ("检测项目", 18),
    ("检测参数", 22),
    ("检测时机", 18),
    ("检验批/取样频率", 28),
    ("检测标准", 22),
]


def generate_testing_plan(
    output_path: str,
    project_info: Dict,
    testing_plan: List[Dict],
    contract_info: Optional[Dict] = None,
    key_notes: Optional[List[str]] = None,
    project_name: str = "",
    sections: Optional[List[Dict]] = None,
    construction_layers: Optional[List[Dict]] = None,  # V4.7
    progress_callback=None  # V6.0.1: 进度回调 (str) → None
) -> str:
    """生成送检计划 Excel（V4.7: 3 Sheet — 封面 + 送检计划 + 工序流程）"""
    # V5.3: 输入校验 — 防止非预期类型导致崩溃
    if not isinstance(testing_plan, list):
        testing_plan = []
    if not isinstance(construction_layers, list):
        construction_layers = []
    if not isinstance(sections, list):
        sections = []
    if not isinstance(key_notes, list):
        key_notes = []
    if not isinstance(project_info, dict):
        project_info = {}

    def _progress(msg):
        if progress_callback:
            try:
                progress_callback(msg)
            except Exception:
                pass

    _progress("生成封面...")
    wb = openpyxl.Workbook()

    # Sheet 1: 封面
    ws = wb.active
    ws.title = "封面"
    _gen_cover(ws, project_info, project_name, key_notes)

    # Sheet 2: 按路段送检计划（16列）
    _progress(f"生成送检计划表 ({len(testing_plan)} 项)...")
    ws2 = wb.create_sheet("按路段送检计划")
    _gen_section_plan(ws2, testing_plan, project_info, project_name)

    # Sheet 3: 施工检测工序流程 (V6.0: 12列)
    if construction_layers:
        _progress(f"生成工序流程表 ({len(construction_layers)} 层)...")
        ws3 = wb.create_sheet("施工检测工序流程")
        _gen_process_flow(ws3, construction_layers, project_info, project_name)

    # Sheet 4 (V6.0 optional): 施工步骤明细表
    procedures = []
    for layer in (construction_layers or []):
        procs = layer.get("procedures", [])
        for proc in procs:
            proc["_section"] = layer.get("section_name", layer.get("section", ""))
            proc["_layer_name"] = layer.get("layer_name", "")
        procedures.extend(procs)
    if procedures:
        _progress(f"生成施工步骤明细 ({len(procedures)} 步)...")
        ws4 = wb.create_sheet("施工步骤明细")
        _gen_procedures_sheet(ws4, procedures, project_info, project_name)

    _progress("保存文件...")
    p = Path(output_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(p))
    return str(p)


# ==================== Sheet 1: 封面 ====================

def _gen_cover(ws, info, name, notes):
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 65

    ws.merge_cells("A1:B1")
    c = ws.cell(row=1, column=1, value="工程材料送检计划表")
    c.font = TITLE_FONT; c.alignment = CENTER

    r = 3
    ws.cell(r, 1, value="项目名称").font = DATA_FONT
    ws.cell(r, 2, value=info.get("project_name") or name).font = DATA_FONT
    r += 1

    for label, key in [("工程地点","location"),("道路等级","road_level"),
        ("道路全长","road_length"),("设计速度","design_speed"),
        ("施工单位","builder"),("监理单位","supervisor"),
        ("检测单位","testing_unit")]:
        v = info.get(key, "")
        if v:
            ws.cell(r, 1, value=label).font = DATA_FONT
            ws.cell(r, 2, value=v).font = DATA_FONT
            r += 1

    r += 1; ws.merge_cells(f"A{r}:B{r}")
    ws.cell(r, 1, value="编制说明：").font = Font(name="微软雅黑", size=10, bold=True)
    r += 1
    ns = ["1. 依据施工图纸、项目合同、省站材料检测指南及 GB55032-2022 编制",
          "2. 检测项目及频率按现行国标、行标、广东省地标执行",
          "3. 见证取样按粤建检协【2015】8号要求执行"]
    if notes: ns += notes
    for n in ns:
        ws.merge_cells(f"A{r}:B{r}"); ws.cell(r,1,value=n).font = Font(name="微软雅黑", size=9); r += 1

    r += 2
    ws.cell(r,1,value="编制日期：").font = DATA_FONT
    ws.cell(r,2,value=datetime.now().strftime("%Y年%m月%d日")).font = DATA_FONT


# ==================== Sheet 2: 按路段送检计划（16列） ====================

def _gen_section_plan(ws, plan, info, name):
    cols = GB55032_COLUMNS
    L = get_column_letter(len(cols))

    for i, (_, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 标题
    r = 1
    ws.merge_cells(f"A{r}:{L}{r}")
    c = ws.cell(r, 1, value=f"{info.get('project_name') or name} — 送检计划（按路段）")
    c.font = TITLE_FONT; c.alignment = CENTER
    ws.row_dimensions[r].height = 32

    # 表头
    r = 2
    for ci, (cn, _) in enumerate(cols, 1):
        c = ws.cell(r, ci, value=cn)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = THIN_BORDER
    ws.row_dimensions[r].height = 26

    # 数据 — 按路段分组
    current_r = 3
    seq = 0
    seen_sections = {}
    section_colors = ["2F5496", "538135", "BF5700", "4E2F9E", "B8860B", "4472C4", "1F4E79"]

    for item in (plan or []):
        sec = item.get("section", "")
        if sec not in seen_sections:
            idx = len(seen_sections) % len(section_colors)
            seen_sections[sec] = idx
        else:
            idx = seen_sections[sec]

        color = section_colors[idx]
        sfill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        seq += 1
        ws.row_dimensions[current_r].height = 32
        dfill = ROW_ODD_FILL if seq % 2 == 0 else ROW_EVEN_FILL

        data = [
            str(seq), sec,
            item.get("road_orientation", ""),
            item.get("lane_count", ""),
            item.get("sub_project",""),
            item.get("sub_sub_project",""),
            item.get("work_item",""),
            item.get("material_name",""), item.get("spec",""),
            item.get("test_item",""), item.get("test_param",""),
            item.get("standard",""), item.get("sampling_method",""),
            item.get("inspection_type",""), item.get("frequency",""),
            item.get("planned_batches",""), item.get("remarks","")
        ]
        for ci, v in enumerate(data, 1):
            c = ws.cell(current_r, ci, value=v)
            c.font = DATA_FONT; c.fill = dfill; c.border = THIN_BORDER
            c.alignment = CENTER if ci in (1, 2, 3, 4, 15, 16) else LEFT
        current_r += 1

    # 合并单元格：相邻行路段(桩号)和路幅相同时合并
    _merge_section_cells(ws, 3, current_r - 1)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{L}{current_r-1}"


def _merge_section_cells(ws, start_row: int, end_row: int):
    """合并相邻行相同路段(桩号)和路幅的单元格（B/C/D列）"""
    if end_row <= start_row:
        return

    col_section = 2  # B列 = 路段(桩号)
    col_road_orientation = 3  # C列 = 路幅
    col_lane_count = 4  # D列 = 车道

    i = start_row
    while i <= end_row:
        sec_val = ws.cell(i, col_section).value
        lr_val = ws.cell(i, col_road_orientation).value
        j = i + 1
        while j <= end_row:
            if (ws.cell(j, col_section).value == sec_val and
                    ws.cell(j, col_road_orientation).value == lr_val):
                j += 1
            else:
                break
        if j - i > 1:
            ws.merge_cells(
                start_row=i, start_column=col_section,
                end_row=j - 1, end_column=col_section
            )
            ws.merge_cells(
                start_row=i, start_column=col_road_orientation,
                end_row=j - 1, end_column=col_road_orientation
            )
            ws.merge_cells(
                start_row=i, start_column=col_lane_count,
                end_row=j - 1, end_column=col_lane_count
            )
        i = j


# ==================== Sheet 3: 施工检测工序流程 (V4.7) ====================

def _gen_process_flow(ws, layers, info, name):
    """V6.0: 施工检测工序流程 Sheet — 12列4级层级（路段→路幅→施工层序→材料/检测）"""
    cols = PROCESS_FLOW_COLS
    L = get_column_letter(len(cols))

    for i, (_, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    r = 1
    ws.merge_cells(f"A{r}:{L}{r}")
    c = ws.cell(r, 1, value=f"{info.get('project_name') or name} — 施工检测工序流程")
    c.font = TITLE_FONT; c.alignment = CENTER
    ws.row_dimensions[r].height = 32

    # Header
    r = 2
    for ci, (cn, _) in enumerate(cols, 1):
        c = ws.cell(r, ci, value=cn)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = THIN_BORDER
    ws.row_dimensions[r].height = 26

    # Group layers by section
    section_colors = ["2F5496", "538135", "BF5700", "4E2F9E", "B8860B", "4472C4", "1F4E79"]
    current_r = 3
    seq = 0

    for layer in (layers or []):
        sec = layer.get("section_name", layer.get("section", ""))
        road_orientation = layer.get("road_orientation", "")
        step = layer.get("step", 1)  # 施工层序（从下到上）
        layer_name = layer.get("layer_name", "")
        thickness = layer.get("thickness", "")
        process = layer.get("construction_process", "")
        key_points = layer.get("key_points", layer.get("construction_key_points", ""))

        # Materials list — V6.0: 每种材料独立一行
        mats = layer.get("materials", [])
        if not mats:
            mats = [{"name": "", "spec": ""}]

        # Tests list
        tests = layer.get("tests", [])

        for m_idx, mat in enumerate(mats):
            mat_name = mat.get("name", mat.get("material_name", "")) if isinstance(mat, dict) else str(mat)
            mat_spec = mat.get("spec", "") if isinstance(mat, dict) else ""

            if not tests:
                seq += 1
                ws.row_dimensions[current_r].height = 24
                dfill = ROW_ODD_FILL if seq % 2 == 0 else ROW_EVEN_FILL
                vals = [
                    sec,  # A: 路段(桩号)
                    road_orientation,  # B: 路幅
                    str(step),  # C: 施工层序
                    f"{layer_name} — {process}" if process else layer_name,  # D
                    key_points,  # E: 施工要点
                    mat_name,  # F: 材料名称
                    mat_spec,  # G: 规格型号
                    "",  # H: 检测项目
                    "",  # I: 检测参数
                    "",  # J: 检测时机
                    "",  # K: 检验批/取样频率
                    "",  # L: 检测标准
                ]
                for ci, v in enumerate(vals, 1):
                    c = ws.cell(current_r, ci, value=v)
                    c.font = DATA_FONT; c.fill = dfill; c.border = THIN_BORDER
                    c.alignment = CENTER if ci in (1, 2, 3) else LEFT
                current_r += 1

            for t_idx, test in enumerate(tests):
                seq += 1
                ws.row_dimensions[current_r].height = 24
                dfill = ROW_ODD_FILL if seq % 2 == 0 else ROW_EVEN_FILL

                vals = [
                    sec,  # A: 路段(桩号)
                    road_orientation,  # B: 路幅
                    str(step),  # C: 施工层序
                    f"{layer_name} — {process}" if process else layer_name,  # D
                    key_points,  # E: 施工要点
                    mat_name,  # F: 材料名称
                    mat_spec,  # G: 规格型号
                    test.get("test_item", ""),  # H
                    test.get("test_param", ""),  # I
                    test.get("timing", ""),  # J
                    test.get("frequency", ""),  # K
                    test.get("standard", ""),  # L
                ]
                for ci, v in enumerate(vals, 1):
                    c = ws.cell(current_r, ci, value=v)
                    c.font = DATA_FONT; c.fill = dfill; c.border = THIN_BORDER
                    c.alignment = CENTER if ci in (1, 2, 3) else LEFT
                current_r += 1

    # V6.0: 合并路段/路幅/层序/层工序 列
    _merge_process_flow_cells_v6(ws, 3, current_r - 1)

    # V6.0: 打印设置 — A3横向, 重复标题行, fitToPage
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 8  # A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:2"  # 重复标题行
    ws.freeze_panes = "A3"


def _merge_process_flow_cells_v6(ws, start_row: int, end_row: int):
    """V6.0: 合并同路段/路幅/层序/层工序的单元格（A-D列）"""
    if end_row <= start_row:
        return

    # A=路段(桩号), B=路幅, C=施工层序, D=施工层/工序, E=施工要点
    cols_to_merge = [1, 2, 3, 4, 5]
    i = start_row
    while i <= end_row:
        sec_val = ws.cell(i, 1).value
        lr_val = ws.cell(i, 2).value
        step_val = ws.cell(i, 3).value
        j = i + 1
        while j <= end_row:
            if (ws.cell(j, 1).value == sec_val and
                ws.cell(j, 2).value == lr_val and
                ws.cell(j, 3).value == step_val):
                j += 1
            else:
                break
        if j - i > 1:
            for col in cols_to_merge:
                ws.merge_cells(
                    start_row=i, start_column=col,
                    end_row=j - 1, end_column=col
                )
        i = j


# ==================== Sheet 4: 施工步骤明细 (V6.0) ====================

PROCEDURE_COLS = [
    ("序号", 5),
    ("路段(桩号)", 16),
    ("所属层位", 18),
    ("步骤序号", 8),
    ("步骤名称", 18),
    ("步骤描述", 36),
    ("施工要点", 28),
    ("适用标准", 22),
    ("施工参数", 30),
]


def _gen_procedures_sheet(ws, procedures, info, name):
    """V6.0: 施工步骤明细表 — 每行一个施工步骤"""
    cols = PROCEDURE_COLS
    L = get_column_letter(len(cols))

    for i, (_, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    r = 1
    ws.merge_cells(f"A{r}:{L}{r}")
    c = ws.cell(r, 1, value=f"{info.get('project_name') or name} — 施工步骤明细")
    c.font = TITLE_FONT; c.alignment = CENTER
    ws.row_dimensions[r].height = 32

    # Header
    r = 2
    for ci, (cn, _) in enumerate(cols, 1):
        c = ws.cell(r, ci, value=cn)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = THIN_BORDER
    ws.row_dimensions[r].height = 26

    current_r = 3
    for idx, proc in enumerate(procedures, 1):
        ws.row_dimensions[current_r].height = 28
        dfill = ROW_ODD_FILL if idx % 2 == 0 else ROW_EVEN_FILL

        params = proc.get("parameters", {})
        param_text = "; ".join(f"{k}: {v}" for k, v in params.items()) if params else ""

        vals = [
            str(idx),
            proc.get("_section", proc.get("section", "")),
            proc.get("_layer_name", proc.get("layer_name", "")),
            str(proc.get("step_order", "")),
            proc.get("step_name", ""),
            proc.get("step_description", ""),
            proc.get("key_points", ""),
            proc.get("applicable_standards", ""),
            param_text,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(current_r, ci, value=v)
            c.font = DATA_FONT; c.fill = dfill; c.border = THIN_BORDER
            c.alignment = CENTER if ci in (1, 4) else LEFT
        current_r += 1

    ws.freeze_panes = "A3"
